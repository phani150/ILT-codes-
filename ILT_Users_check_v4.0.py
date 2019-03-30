###################################################################################################################################################
#   File Name       : Users Data Dump Analysis.py
#   Description     : Users  Data Dump Utility 
#   Author          : D Phanindra(phaindra.d@unilever.com)
#   Version History :
#   Date            Version     Change Description
#   21-Sep-2018     3.0         Initial Draft
####################################################################################################################################################

##Importing the required modules 

import os
import sys
import shutil
import xlrd
import openpyxl
from openpyxl.workbook.workbook import Workbook # as _Workbook
import csv
import pandas as pd
import numpy as np
from pprint import pprint 
from datetime import datetime
from collections import defaultdict

status=0
SourceFileFound=0
global alerts_dict,excess_count_dict
alerts_dict={}
excess_count_dict={}
SourceDict={}
SourceDict2={}
norms_dict={}
src_dummy={}
cwd = os.getcwd()

print("************************** OPTIONS MENU *********************************")
print("1. Generate Reconciliation Results for Norms Issue")
print("2. Generate Reconciliation Results for Chronic stock service Issue")
print("3. Generate Reconciliation Results for In-Tarnsit Issue")
print("4. Generate Reconciliation Results for Chronic excess stock Issue")
print("5. Exit")
while(status != 1):
    inputValue = input("\nPlease enter your choice : ")
    if(inputValue not in ['1','2','3','4','5']):
        print("Invalid Input!!!")
        status=0
    elif(inputValue == '5'):
        sys.exit()
    else:
        status=1        
if(inputValue == '1'):
    outputFileName = "dummy_Issues1.txt"
    outputFile = os.path.join(cwd,outputFileName)
    outputFileName2 = "Req_rr_weeks.txt"
    outputFile2 = os.path.join(cwd,outputFileName2)
    outputFileName3 = "Issues1_without_dup.txt"
    outputFile3 = os.path.join(cwd,outputFileName3)
    outputFileName4 = "Norms_Issues.txt"
    outputFile4 = os.path.join(cwd,outputFileName4)
    outputDirectoryName = "Norms Issue Files"
    outputDirectory = os.path.join(cwd,outputDirectoryName)
    if os.path.exists(outputDirectory):
        shutil.rmtree(outputDirectory)
    if not os.path.exists(outputDirectory):
        os.makedirs(outputDirectory)
    os.chdir(cwd)
    
    ##Reading Config.txt file 
    for file in os.listdir(cwd):
        if(file.endswith(".txt") and "Config." in file):
            configFound = 1
            f = open(file,"r")
            rowR = f.readlines()
            for line in rowR:
                if(line.strip() == ""):
                    continue
                else:
                    lineData = line.split("=")
                    if(lineData[0] == "Users Data Dump"):
                        if(lineData[1].strip() == ""):
                            print("Config file error !!! Source_DIS file name source File Namenot specified in Config.txt file !!!")
                            print("\nAborting the program !!!")
                            input("Press ENTER key to continue ...")
                            sys.exit()
                        else:
                            SourceFile = lineData[1].strip()       
            break
        else:
            configFound = 0
    if(configFound == 0):
        print("\nConfig.txt file not found in the directory!!! Aborting the program!!!\n")
        input("Press ENTER key to continue ...")
        sys.exit()
    
    ls=[]
    for file in os.listdir(cwd):
        if(file ==  SourceFile):
            SourceFileFound=1
            print("\n Processing Data dump Report  file ...")
            wb_list=xlrd.open_workbook(file).sheet_names()
            worksheet= xlrd.open_workbook(file)
            norms_alert=worksheet.sheet_by_name("Users input")
            data_dump= worksheet.sheet_by_name("Data Dump")
            ##Reding the i/p alerts and building the alerts_dict
            for data in range(1,norms_alert.nrows):
                usr_checks= str(norms_alert.cell(data,0).value).strip()
                inputs= int(float(str(norms_alert.cell(data,1).value).strip()))
                alerts_dict.update({usr_checks:inputs})
                
            global rr_weeks,max_threshold_norms,min_threshold_norms,RR_greater_norms,RR_range,wk_penetration,mstn_threshold,intransit_norms,MSTN_upper_threshold
            
            intransit_norms=int(alerts_dict["Intransit Norms threshold"])
            MSTN_upper_threshold=int(alerts_dict["MSTN Upper Threshold"])
            wk_penetration=int(alerts_dict['% Wk penetration below threshold'])
            mstn_threshold=int(alerts_dict['MSTN Low Threshold'])
            rr_weeks=int(alerts_dict['Duration for which Norm RR to be evaluated'])
            max_threshold_norms=int(alerts_dict['Max Threshold Norms (in Wks)'])
            min_threshold_norms=int(alerts_dict['Min Threshold Norms (in Wks)'])
            RR_greater=int(alerts_dict['Evaluate packs with RR greater than(>)'])
            RR_range= int(alerts_dict["Evaluate packs with RR in the given range"])
            
            ##Reading the Data Dump sheet and building the  SourceDict
            for row in range(1,data_dump.nrows):
                cust_code= int(float(str(data_dump.cell(row,1).value).strip()))
                cust_name=str(data_dump.cell(row,2).value).strip()
                sku_map=int(float(str(data_dump.cell(row,3).value).strip()))
                unique=str(cust_code)+str(sku_map)
                sku_desc= str(data_dump.cell(row,4).value).strip()
                phy_norms=str(data_dump.cell(row,6).value).strip()
                if (phy_norms == ''):
                    phy_norms=phy_norms.replace(",","")
                thr_transit_WC=str(data_dump.cell(row,7).value).strip()
                transit_norm= str(data_dump.cell(row,8).value).strip()
                if (transit_norm == ''):
                    transit_norm=transit_norm.replace('','0.0')
                phy_stock=str(data_dump.cell(row,9).value).strip()
                transit_stock= str(data_dump.cell(row,10).value).strip()
                if (transit_stock == ''):
                    transit_stock=transit_stock.replace('','0.0')
                rr=str(data_dump.cell(row,11).value).strip()
                tot_stock=str(float(phy_stock)+float(transit_stock))
                tot_norms=str(float(phy_norms)+float(transit_norm))
                target_st=str(data_dump.cell(row,12).value).strip()
                week= int(float(str(data_dump.cell(row,5).value).strip()))
                ls.append(week)
                phy_ratio="0.0"
                if (phy_norms != "0.0"):
                     phy_ratio=str((float(phy_stock)/float(phy_norms)*100))
                else:
                    phy_ratio="0.0"
                totals_ratio="0.0"
                if (tot_norms != "0.0"):
                    totals_ratio=str((float(tot_stock)/float(tot_norms)*100))
                else:
                    totals_ratio="0.0"
                key=str(cust_code)+"|"+str(sku_map)+"|"+str(week)
                ValueString=str(unique)+"|"+str(cust_name)+"|"+str(sku_desc)+"|"+str(phy_norms)+"|"+str(thr_transit_WC)+"|"+str(transit_norm)+"|"+str(phy_stock)+"|"+str(transit_stock)+"|"+str(rr)+"|"+str(target_st)
                ValueString=ValueString+"|"+str(phy_ratio)+"|"+str(totals_ratio)
                src_dummy.update({key:ValueString})    
            week_list=sorted(list(set(ls)))
            max_week=int(max(week_list))
            start_week=int(max_week) - int(rr_weeks)
            
            ##Getting the required rr_weeks data
            for src_key in src_dummy:
                src_split=src_key.split("|")
                cust_code=src_split[0].strip()
                sku_map=src_split[1].strip()
                week=src_split[2].strip()
                valueSplit=src_dummy[src_key].split("|")
                rr=float(valueSplit[8].strip())
                if (int(start_week)<int(week)<=int(max_week)):
                    if (float(rr) > float(RR_greater)) and (float(rr) < float(RR_range)):
                        key=str(cust_code)+"|"+str(sku_map)+"|"+str(week)
                        ValueString=src_dummy[src_key]
                        SourceDict.update({key:ValueString})            
            print("Processing of Data dump file completed !!!")
            break
        else:
            SourceFileFound=0

      
    f2= open(outputFile2,"w+")
    headerString2 = 'Customer code|SKU MAP|week|Unique|Customer name|SKU DESC|Theoritical Norm qty|Theoritical In Transit in WC|Theoritical In Transit in cs|'
    headerString2=headerString2 + "Actual Stock on hand|Actual In transit|Define CRR/RPP|Target Stock|Physical Ratio|Totals Ratio\n"
    f2.write(headerString2)
    f3= open(outputFile3,"w+")
    headerString3= 'Customer code|SKU MAP|'
    headerString3=headerString3 + "Physical sum|RR sum|Norms Ratio|Condition\n"
    f3.write(headerString3)

    f4= open(outputFile4,"w+")
    headerString4= 'Customer code|SKU MAP|'
    headerString4=headerString4 + "Physical sum|RR sum|Norms Ratio|Condition\n"
    f4.write(headerString4)
    
    
    ##Reading the SourceDict and cal norms
    key_list=[]
    phy_list=[]
    rr_list=[] 
    for key in SourceDict:
        keySplit=key.split("|")
        cust_code=keySplit[0].strip()
        sku_map=keySplit[1].strip()
        week=keySplit[2].strip()
        valueSplit=SourceDict[key].split("|")
        phy_norms=float(valueSplit[3].strip())
        rr=float(valueSplit[8].strip())
        outputString2 = key +'|'+SourceDict[key]+ '\n'
        f2.write(outputString2)
    f2.close()
    
    # Load data from Req_rr_weeks.txt file
    data = pd.read_table(outputFile2,sep="|",encoding = "ISO-8859-1",usecols =["Customer code","SKU MAP",'Theoritical Norm qty','Define CRR/RPP'])
    df1=data.groupby(["Customer code","SKU MAP"])[['Theoritical Norm qty','Define CRR/RPP']].sum()
    
    # converting dtypes using astype 
    df1['Theoritical Norm qty']= df1['Theoritical Norm qty'].astype(int)
    df1['Define CRR/RPP']= df1['Define CRR/RPP'].astype(int)
    df1 = df1[(df1[['Theoritical Norm qty','Define CRR/RPP']] != 0).all(axis=1)]
    df1['Ratio']=df1['Theoritical Norm qty']/df1['Define CRR/RPP']
    df1.to_csv(outputFileName,sep="|")

    with open(outputFile,"r") as remv_dup:
        for line in remv_dup:
            if('Customer code' in line or line.strip() == ''):
                continue
            splitData = line.split('|')
            cust_code= splitData[0].strip()
            sku_map= splitData[1].strip()
            phy_sum= float(splitData[2].strip())
            rr_sum=float(splitData[3].strip())
            ratio=float(splitData[4].strip())
            
            norm_key=str(cust_code)+"|"+str(sku_map)
            norm_value=str(phy_sum)+"|"+str(rr_sum)+"|"+str(ratio)
            norms_dict.update({norm_key:norm_value})
            
    for key in norms_dict:
        valueSplit=norms_dict[key].split("|")
        ratio=float(valueSplit[2].strip())
        if float(ratio) < float(min_threshold_norms):
            norms="Increased norms"
            outputString3=key+"|"+norms_dict[key]+"|"+str(norms)+"\n" 
        if float(ratio) > float(max_threshold_norms):
            norms="Decreased norms"
            outputString3=key+"|"+norms_dict[key]+"|"+str(norms)+"\n"     
        f3.write(outputString3)
    f3.close()

    final_norm_dict={}
    with open(outputFile3,"r") as final_norm:
        for line in final_norm:
            if('Customer code' in line or line.strip() == ''):
                continue
            splitData = line.split('|')
            cust_code= splitData[0].strip()
            sku_map= splitData[1].strip()
            phy_sum= float(splitData[2].strip())
            rr_sum=float(splitData[3].strip())
            ratio=float(splitData[4].strip())
            norms=splitData[5].strip()
            final_key=str(cust_code)+"|"+str(sku_map)
            final_value=str(phy_sum)+"|"+str(rr_sum)+"|"+str(ratio)+"|"+str(norms)
            final_norm_dict.update({final_key:final_value})

    for final_key in final_norm_dict:
        outputString4=final_key+"|"+ final_norm_dict[final_key] +"\n"
        f4.write(outputString4)
    f4.close()        

    # Opening the Reports data
    xfile = openpyxl.load_workbook("Reports Data.xlsx")
    # Add sheet to  the Reports data sheet 
    xfile.create_sheet('Norms Issue')
    ws = xfile['Norms Issue']   
    with open(outputFileName4) as tab_file:
        tab_reader = csv.reader(tab_file, delimiter='|')
        for rowData in tab_reader:
            ws.append(rowData)
        xfile.save('Reports Data.xlsx')
    tab_file.close()   
    
    #Generating  the Norms issue  File directory and removing 
    shutil.copy(os.path.join(cwd,outputFileName),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName))
    shutil.copy(os.path.join(cwd,outputFileName2),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName2))
    shutil.copy(os.path.join(cwd,outputFileName3),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName3))
    shutil.copy(os.path.join(cwd,outputFileName4),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName4))
    
    
if(inputValue == '2'):
    outputFileName = "Service_issue.txt"
    outputFile = os.path.join(cwd,outputFileName)
    outputFileName2= "Req_Service_issue.csv"
    outputFile2 = os.path.join(cwd,outputFileName2)
    outputDirectoryName = "Chronic service Files"
    outputDirectory = os.path.join(cwd,outputDirectoryName)
    if os.path.exists(outputDirectory):
        shutil.rmtree(outputDirectory)
    if not os.path.exists(outputDirectory):
        os.makedirs(outputDirectory)
    os.chdir(cwd)
    ##Reading Config.txt file 
    for file in os.listdir(cwd):
        if(file.endswith(".txt") and "Config." in file):
            configFound = 1
            f = open(file,"r")
            rowR = f.readlines()
            for line in rowR:
                if(line.strip() == ""):
                    continue
                else:
                    lineData = line.split("=")
                    if(lineData[0] == "Users Data Dump"):
                        if(lineData[1].strip() == ""):
                            print("Config file error !!! Source_DIS file name source File Namenot specified in Config.txt file !!!")
                            print("\nAborting the program !!!")
                            input("Press ENTER key to continue ...")
                            sys.exit()
                        else:
                            SourceFile = lineData[1].strip()       
            break
        else:
            configFound = 0
    if(configFound == 0):
        print("\nConfig.txt file not found in the directory!!! Aborting the program!!!\n")
        input("Press ENTER key to continue ...")
        sys.exit()
        
    ls=[]
    for file in os.listdir(cwd):
        if(file ==  SourceFile):
            SourceFileFound=1
            print("\n Processing Based Reports  file ...")
            wb_list=xlrd.open_workbook(file).sheet_names()
            worksheet= xlrd.open_workbook(file)
            norms_alert=worksheet.sheet_by_name("Users input")
            data_dump= worksheet.sheet_by_name("Data Dump")
            
            ##Reding the i/p alerts and building the alerts_dict
            for data in range(1,norms_alert.nrows):
                usr_checks= str(norms_alert.cell(data,0).value).strip()
                inputs= int(float(str(norms_alert.cell(data,1).value).strip()))
                alerts_dict.update({usr_checks:inputs})
            
            intransit_norms=int(alerts_dict["Intransit Norms threshold"])
            MSTN_upper_threshold=int(alerts_dict["MSTN Upper Threshold"])
            wk_penetration=int(alerts_dict['% Wk penetration below threshold'])
            mstn_threshold=int(alerts_dict['MSTN Low Threshold'])
            rr_weeks=int(alerts_dict['Duration for which Norm RR to be evaluated'])
            max_threshold_norms=int(alerts_dict['Max Threshold Norms (in Wks)'])
            min_threshold_norms=int(alerts_dict['Min Threshold Norms (in Wks)'])

            ##Reading the Data Dump sheet
            for row in range(1,data_dump.nrows):
                cust_code= int(float(str(data_dump.cell(row,1).value).strip()))
                cust_name=str(data_dump.cell(row,2).value).strip()
                sku_map=int(float(str(data_dump.cell(row,3).value).strip()))
                unique=str(cust_code)+str(sku_map)
                sku_desc= str(data_dump.cell(row,4).value).strip()
                week= str(data_dump.cell(row,5).value).lstrip('0')    
                phy_norms=str(data_dump.cell(row,6).value).strip()
                if (phy_norms == ''):
                    phy_norms=phy_norms.replace('','0.0')
                thr_transit_WC=str(data_dump.cell(row,7).value).strip()
                transit_norms= str(data_dump.cell(row,8).value).strip()
                if (transit_norms == ''):
                    transit_norms=transit_norms.replace('','0.0')
                phy_stock=str(data_dump.cell(row,9).value).strip()
                if (phy_stock == ''):
                    phy_stock=phy_stock.replace('','0.0')
                transit_stock= str(data_dump.cell(row,10).value).strip()
                if (transit_stock == ''):
                    transit_stock=transit_stock.replace('','0.0')
                rr=str(data_dump.cell(row,11).value).strip()
                target_st=str(data_dump.cell(row,12).value).strip()
                tot_stock=str(float(phy_stock)+float(transit_stock))
                tot_norms=str(float(phy_norms)+float(transit_norms))
                phy_ratio="0.0"
                if (phy_norms != "0.0"):
                     phy_ratio=str((float(phy_stock)/float(phy_norms)*100))
                else:
                    phy_ratio="0.0"
                totals_ratio="0.0"
                if (tot_norms != "0.0"):
                    totals_ratio=str((float(tot_stock)/float(tot_norms)*100))
                else:
                    totals_ratio="0.0"
                key=str(cust_code)+"|"+str(sku_map)+"|"+str(week)
                ValueString=str(unique)+"|"+str(cust_name)+"|"+str(sku_desc)+"|"+str(phy_norms)+"|"+str(thr_transit_WC)+"|"+str(transit_norms)+"|"+str(phy_stock)+"|"+str(transit_stock)+"|"+str(rr)+"|"+str(target_st)
                ValueString=ValueString+"|"+str(phy_ratio)+"|"+str(totals_ratio)
                SourceDict.update({key:ValueString})
            print("Processing of Data Dump file completed !!!")
            break
        else:
            SourceFileFound=0
   
    f= open(outputFile,"w+")
    headerString = 'Customer code|SKU MAP|week|Unique|Customer name|SKU DESC|Theoritical Norm qty|Theoritical In Transit in WC|Theoritical In Transit in cs|'
    headerString=headerString + "Actual Stock on hand|Actual In transit| Define CRR/RPP|Target Stock|Physical Ratio|Totals Ratio\n"
    f.write(headerString)

    for key in SourceDict:  
        outputString=key+"|"+SourceDict[key]+"\n"
        f.write(outputString)
    f.close()
    
    # Load data from csv file
    data = pd.read_table("Service_issue.txt",sep="|",encoding = "ISO-8859-1",usecols =["Customer code","SKU MAP","Unique","week","Customer name","SKU DESC","Physical Ratio","Totals Ratio"])
    df1=data.groupby(["Customer code","SKU MAP","Unique","Customer name","SKU DESC"]).agg({"week":[max],"Unique":'count'})
    df2=data[(data['Physical Ratio'] < mstn_threshold) & (data['Totals Ratio'] < mstn_threshold)].groupby(["Customer code","SKU MAP","Unique","Customer name","SKU DESC"]).agg({"week":[max],"Unique":'count'})
    df2["week"]=df1["week"]
    df2["WK_Violation_ratio"]=(df2["Unique"]/df1["Unique"])*100
    finaldf=df2[(df2.WK_Violation_ratio > wk_penetration)]
    finaldf.to_csv(outputFileName2,sep="|")
    
    # Opening the Reports data
    xfile = openpyxl.load_workbook("Reports Data.xlsx")
    # Add sheet to  the Reports data sheet 
    xfile.create_sheet('Chronic service Issue')
    ws = xfile['Chronic service Issue']   
    with open("Req_Service_issue.csv") as tab_file:
        tab_reader = csv.reader(tab_file, delimiter='|')
        for rowData in tab_reader:
            ws.append(rowData)
        xfile.save('Reports Data.xlsx')
    tab_file.close()  
    
    #Generating  the Chronic service Files directory and removing 
    shutil.copy(os.path.join(cwd,outputFileName),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName))
    shutil.copy(os.path.join(cwd,outputFileName2),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName2))


if(inputValue == '3'):
    outputFileName = "In transit_issue.txt"
    outputFile = os.path.join(cwd,outputFileName)
    outputFileName2 = "req_transit.txt"
    outputFile2 = os.path.join(cwd,outputFileName2)
    outputDirectoryName = "In-trnsit service File"
    outputDirectory = os.path.join(cwd,outputDirectoryName)
    if os.path.exists(outputDirectory):
        shutil.rmtree(outputDirectory)
    if not os.path.exists(outputDirectory):
        os.makedirs(outputDirectory)
    os.chdir(cwd)
    ##Reading Config.txt file 
    for file in os.listdir(cwd):
        if(file.endswith(".txt") and "Config." in file):
            configFound = 1
            f = open(file,"r")
            rowR = f.readlines()
            for line in rowR:
                if(line.strip() == ""):
                    continue
                else:
                    lineData = line.split("=")
                    if(lineData[0] == "Users Data Dump"):
                        if(lineData[1].strip() == ""):
                            print("Config file error !!! Source_DIS file name source File Name not specified in Config.txt file !!!")
                            print("\nAborting the program !!!")
                            input("Press ENTER key to continue ...")
                            sys.exit()
                        else:
                            SourceFile = lineData[1].strip()       
            break
        else:
            configFound = 0
    if(configFound == 0):
        print("\nConfig.txt file not found in the directory!!! Aborting the program!!!\n")
        input("Press ENTER key to continue ...")
        sys.exit()

    for file in os.listdir(cwd):
        if(file ==  SourceFile):
            SourceFileFound=1
            print("\n Processing Based Reports  file ...")
            wb_list=xlrd.open_workbook(file).sheet_names()
            worksheet= xlrd.open_workbook(file)
            norms_alert=worksheet.sheet_by_name("Users input")
            data_dump= worksheet.sheet_by_name("Data Dump")
            
            ##Reding the i/p alerts and building the alerts_dict
            for data in range(1,norms_alert.nrows):
                usr_checks= str(norms_alert.cell(data,0).value).strip()
                inputs= int(float(str(norms_alert.cell(data,1).value).strip()))
                alerts_dict.update({usr_checks:inputs})
            
            intransit_norms=int(alerts_dict["Intransit Norms threshold"])
            MSTN_upper_threshold=int(alerts_dict["MSTN Upper Threshold"])
            wk_penetration=int(alerts_dict['% Wk penetration below threshold'])
            mstn_threshold=int(alerts_dict['MSTN Low Threshold'])
            rr_weeks=int(alerts_dict['Duration for which Norm RR to be evaluated'])
            max_threshold_norms=int(alerts_dict['Max Threshold Norms (in Wks)'])
            min_threshold_norms=int(alerts_dict['Min Threshold Norms (in Wks)'])
         
            ##Reading the Data Dump sheet
            for row in range(1,data_dump.nrows):
                cust_code= int(float(str(data_dump.cell(row,1).value).strip()))
                cust_name=str(data_dump.cell(row,2).value).strip()
                sku_map=int(float(str(data_dump.cell(row,3).value).strip()))
                sku_desc= str(data_dump.cell(row,4).value).strip()
                unique=str(cust_code)+str(sku_map)
                week= str(data_dump.cell(row,5).value).lstrip('0')    
                phy_norms=str(data_dump.cell(row,6).value).strip()
                if (phy_norms == ''):
                    phy_norms=phy_norms.replace('','0.0')
                thr_transit_WC=str(data_dump.cell(row,7).value).strip()
                transit_norms= str(data_dump.cell(row,8).value).strip()
                if (transit_norms == ''):
                    transit_norms=transit_norms.replace('','0.0')
                phy_stock=str(data_dump.cell(row,9).value).strip()
                if (phy_stock == ''):
                    phy_stock=phy_stock.replace('','0.0')
                transit_stock= str(data_dump.cell(row,10).value).strip()
                if (transit_stock == ''):
                    transit_stock=transit_stock.replace('','0.0')
                rr=str(data_dump.cell(row,11).value).strip()
                target_st=str(data_dump.cell(row,12).value).strip()
                tot_stock=str(float(phy_stock)+float(transit_stock))
                tot_norms=str(float(phy_norms)+float(transit_norms))
                phy_ratio="0.0"
                if (phy_norms != "0.0"):
                     phy_ratio=str((float(phy_stock)/float(phy_norms)*100))
                else:
                    phy_ratio="0.0"
                totals_ratio="0.0"
                if (tot_norms != "0.0"):
                    totals_ratio=str((float(tot_stock)/float(tot_norms)*100))
                else:
                    totals_ratio="0.0"
                key=str(cust_code)+"|"+str(sku_map)+"|"+str(week)
                ValueString=str(unique)+"|"+str(cust_name)+"|"+str(sku_desc)+"|"+str(phy_norms)+"|"+str(thr_transit_WC)+"|"+str(transit_norms)+"|"+str(phy_stock)+"|"+str(transit_stock)+"|"+str(rr)+"|"+str(target_st)
                ValueString=ValueString+"|"+str(phy_ratio)+"|"+str(totals_ratio)
                SourceDict.update({key:ValueString})
            print("Processing of Data Dump file completed !!!")
            break
        else:
            SourceFileFound=0
   
   
    f= open(outputFile,"w+")
    headerString = 'Customer code|SKU MAP|week|Unique|Customer name|SKU DESC|Theoritical Norm qty|Theoritical In Transit in WC|Theoritical In Transit in cs|'
    headerString=headerString + "Actual Stock on hand|Actual In transit| Define CRR/RPP|Target Stock|Physical Ratio|Transit Ratio\n"
    f.write(headerString)
    
    for key in SourceDict:  
        outputString=key+"|"+SourceDict[key]+"\n"
        f.write(outputString)
    f.close()

    # Load data from txt file
    data = pd.read_table("In transit_issue.txt",sep="|",encoding = "ISO-8859-1",usecols =["Customer code","SKU MAP","Unique","week","Customer name","SKU DESC","Physical Ratio","Transit Ratio"])
    df1=data.groupby(["Customer code","SKU MAP","Unique","Customer name","SKU DESC"]).agg({"week":[max],"Unique":'count'})
    df2=data[(data["Transit Ratio"] > MSTN_upper_threshold) & (data["Physical Ratio"]< MSTN_upper_threshold)].groupby(["Customer code","SKU MAP","Unique","Customer name","SKU DESC"]).agg({"week":[max],"Unique":'count'})
    df2['week']=df1['week']
    df2["WK_Violation_ratio"]=(df2["Unique"]/df1["Unique"])*100
    finaldf=df2[(df2.WK_Violation_ratio > wk_penetration)]
    finaldf.to_csv(outputFileName2,sep="|")

    # Opening the Reports data
    xfile = openpyxl.load_workbook("Reports Data.xlsx")
    # Add sheet to  the Reports data sheet 
    xfile.create_sheet("In Transit stock Issue")
    ws = xfile["In Transit stock Issue"]   
    with open("req_transit.txt") as tab_file:
        tab_reader = csv.reader(tab_file, delimiter='|')
        for rowData in tab_reader:
            ws.append(rowData)
        xfile.save('Reports Data.xlsx')
    tab_file.close()   
    
        
    #Generating  the In-trnsit service File directory and removing 
    shutil.copy(os.path.join(cwd,outputFileName),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName))
    shutil.copy(os.path.join(cwd,outputFileName2),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName2))

    
if(inputValue == '4'):
    outputFileName = "dummy chronic excess stock.txt"
    outputFile = os.path.join(cwd,outputFileName)
    outputFileName2 = "req_excess.csv"
    outputFile2 = os.path.join(cwd,outputFileName2)
    outputDirectoryName = "Chronic excess issue Files"
    outputDirectory = os.path.join(cwd,outputDirectoryName)
    if os.path.exists(outputDirectory):
        shutil.rmtree(outputDirectory)
    if not os.path.exists(outputDirectory):
        os.makedirs(outputDirectory)
    os.chdir(cwd)
    ##Reading Config.txt file 
    for file in os.listdir(cwd):
        if(file.endswith(".txt") and "Config." in file):
            configFound = 1
            f = open(file,"r")
            rowR = f.readlines()
            for line in rowR:
                if(line.strip() == ""):
                    continue
                else:
                    lineData = line.split("=")
                    if(lineData[0] == "Users Data Dump"):
                        if(lineData[1].strip() == ""):
                            print("Config file error !!! Source_DIS file name source File Namenot specified in Config.txt file !!!")
                            print("\nAborting the program !!!")
                            input("Press ENTER key to continue ...")
                            sys.exit()
                        else:
                            SourceFile = lineData[1].strip()       
            break
        else:
            configFound = 0
    if(configFound == 0):
        print("\nConfig.txt file not found in the directory!!! Aborting the program!!!\n")
        input("Press ENTER key to continue ...")
        sys.exit()

    for file in os.listdir(cwd):
        if(file ==  SourceFile):
            SourceFileFound=1
            print("\n Processing Based Reports  file ...")
            wb_list=xlrd.open_workbook(file).sheet_names()
            worksheet= xlrd.open_workbook(file)
            norms_alert=worksheet.sheet_by_name("Users input")
            data_dump= worksheet.sheet_by_name("Data Dump")
            
            ##Reding the i/p alerts and building the alerts_dict
            for data in range(1,norms_alert.nrows):
                usr_checks= str(norms_alert.cell(data,0).value).strip()
                inputs= int(float(str(norms_alert.cell(data,1).value).strip()))
                alerts_dict.update({usr_checks:inputs})
            
            intransit_norms=int(alerts_dict["Intransit Norms threshold"])
            MSTN_upper_threshold=int(alerts_dict["MSTN Upper Threshold"])
            wk_penetration=int(alerts_dict['% Wk penetration below threshold'])
            mstn_threshold=int(alerts_dict['MSTN Low Threshold'])
            rr_weeks=int(alerts_dict['Duration for which Norm RR to be evaluated'])
            max_threshold_norms=int(alerts_dict['Max Threshold Norms (in Wks)'])
            min_threshold_norms=int(alerts_dict['Min Threshold Norms (in Wks)'])
         
            ##Reading the Data Dump sheet
            for row in range(1,data_dump.nrows):
                cust_code= int(float(str(data_dump.cell(row,1).value).strip()))
                cust_name=str(data_dump.cell(row,2).value).strip()
                sku_map=int(float(str(data_dump.cell(row,3).value).strip()))
                unique=str(cust_code)+str(sku_map)
                sku_desc= str(data_dump.cell(row,4).value).strip()
                week= str(data_dump.cell(row,5).value).lstrip('0')    
                phy_norms=str(data_dump.cell(row,6).value).strip()
                if (phy_norms == ''):
                    phy_norms=phy_norms.replace('','0.0')
                thr_transit_WC=str(data_dump.cell(row,7).value).strip()
                transit_norms= str(data_dump.cell(row,8).value).strip()
                if (transit_norms == ''):
                    transit_norms=transit_norms.replace('','0.0')
                phy_stock=str(data_dump.cell(row,9).value).strip()
                if (phy_stock == ''):
                    print (phy_stock)
                    phy_stock=phy_stock.replace('','0.0')
                transit_stock= str(data_dump.cell(row,10).value).strip()
                rr=str(data_dump.cell(row,11).value).strip()
                target_st=str(data_dump.cell(row,12).value).strip()
                tot_stock=str(float(phy_stock)+float(transit_stock))
                tot_norms=str(float(phy_norms)+float(transit_norms))
                phy_ratio="0.0"
                if (phy_norms != "0.0"):
                     phy_ratio=str((float(phy_stock)/float(phy_norms)*100))
                else:
                    phy_ratio="0.0"
                totals_ratio="0.0"
                if (tot_norms != "0.0"):
                    totals_ratio=str((float(tot_stock)/float(tot_norms)*100))
                else:
                    totals_ratio="0.0"
                key=str(cust_code)+"|"+str(sku_map)+"|"+str(week)
                ValueString=str(unique)+"|"+str(cust_name)+"|"+str(sku_desc)+"|"+str(phy_norms)+"|"+str(thr_transit_WC)+"|"+str(transit_norms)+"|"+str(phy_stock)+"|"+str(transit_stock)+"|"+str(rr)+"|"+str(target_st)
                ValueString=ValueString+"|"+str(phy_ratio)+"|"+str(totals_ratio)
                SourceDict.update({key:ValueString})
                
            print("Processing of Data Dump file completed !!!")
            break
        else:
            SourceFileFound=0
   
   
    f= open(outputFile,"w+")
    headerString = 'Customer code|SKU MAP|week|Unique|Customer name|SKU DESC|Theoritical Norm qty|Theoritical In Transit in WC|Theoritical In Transit in cs|'
    headerString=headerString + "Actual Stock on hand|Actual In transit| Define CRR/RPP|Target Stock|Physical Ratio|Totals Ratio\n"
    f.write(headerString)
    
    for key in SourceDict:  
        outputString=key+"|"+SourceDict[key]+"\n"
        f.write(outputString)
    f.close()
    
    # Load data from csv file
    data = pd.read_table("dummy chronic excess stock.txt",sep="|",encoding = "ISO-8859-1",usecols =["Customer code","SKU MAP","Unique","week","Customer name","SKU DESC","Physical Ratio","Totals Ratio"])
    df1=data.groupby(["Customer code","SKU MAP","Unique","Customer name","SKU DESC"]).agg({"week":[max],"Unique":'count'})
    df2=data[(data['Physical Ratio'] > MSTN_upper_threshold) & (data['Totals Ratio']>MSTN_upper_threshold)].groupby(["Customer code","SKU MAP","Unique","Customer name","SKU DESC"]).agg({"week":[max],"Unique":'count'})
    df2["week"]=df1["week"]
    df2["WK_Violation_ratio"]=(df2["Unique"]/df1["Unique"])*100
    finaldf=df2[(df2.WK_Violation_ratio > wk_penetration)]
    finaldf.to_csv(outputFileName2,sep="|")


    # Opening the Reports data  
    xfile = openpyxl.load_workbook("Reports Data.xlsx")
    # Add sheet to  the Reports data sheet 
    xfile.create_sheet('Chronic excess stock')
    ws = xfile["Chronic excess stock"]   
    with open('req_excess.csv') as tab_file:
        tab_reader = csv.reader(tab_file, delimiter='|')
        for rowData in tab_reader:
            ws.append(rowData)
        xfile.save('Reports Data.xlsx')
    tab_file.close()   
    
    #Generating  the Chronic excess issue  File directory and removing 
    shutil.copy(os.path.join(cwd,outputFileName),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName))
    shutil.copy(os.path.join(cwd,outputFileName2),os.path.join(cwd,outputDirectoryName))
    os.remove(os.path.join(cwd,outputFileName2))
   

#Process end time 
end = datetime.now()
print("Processing end time : "+str(end))
print("\n")
print("Please check Output Files Directory for results ...")
input("Press ENTER key to continue ...")
